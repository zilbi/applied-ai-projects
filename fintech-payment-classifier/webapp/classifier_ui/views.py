from __future__ import annotations
import json
import io
from pathlib import Path
from django.conf import settings
from django.db.models import Q
from django.http import FileResponse, Http404, HttpResponse
from django.shortcuts import get_object_or_404,redirect,render
from django.utils import timezone
from .forms import ManualOverrideForm,RunUploadForm
from .models import ClassificationRun,CompanyResult
from .services import save_upload,start_run

def home(request): return render(request,'classifier_ui/home.html',{'runs':ClassificationRun.objects.all()[:15],'count':ClassificationRun.objects.count()})
def new_run(request):
    if request.method=='POST':
        form=RunUploadForm(request.POST,request.FILES)
        if form.is_valid():
            file=form.cleaned_data['file']; path=save_upload(file)
            run=ClassificationRun.objects.create(original_filename=Path(file.name).name,stored_input_path=str(path),external_network_enabled=True,current_stage='created')
            start_run(run); return redirect('classifier_ui:run_detail',run_id=run.id)
    else: form=RunUploadForm()
    return render(request,'classifier_ui/new_run.html',{'form':form})
def run_detail(request,run_id): return render(request,'classifier_ui/run_detail.html',{'run':get_object_or_404(ClassificationRun,id=run_id)})
def company_list(request,run_id):
    run=get_object_or_404(ClassificationRun,id=run_id); rows=run.companies.all()
    q=request.GET.get('q',''); cls=request.GET.get('class',''); decision=request.GET.get('decision',''); site=request.GET.get('site',''); agree=request.GET.get('agree',''); confidence=request.GET.get('min_confidence','')
    if q: rows=rows.filter(Q(canonical_name__icontains=q)|Q(inn__icontains=q))
    if cls: rows=rows.filter(final_class=cls)
    if decision: rows=rows.filter(decision_status=decision)
    if site: rows=rows.filter(site_status=site)
    if agree in {'yes','no'}: rows=rows.filter(models_agree=agree=='yes')
    if confidence:
        try: rows=rows.filter(final_confidence__gte=float(confidence))
        except ValueError: pass
    order=request.GET.get('sort','-final_confidence'); allowed={'final_confidence','-final_confidence','operation_count','-operation_count','final_class','canonical_name','-canonical_name'}; rows=rows.order_by(order if order in allowed else '-final_confidence')
    from django.core.paginator import Paginator
    return render(request,'classifier_ui/company_list.html',{'run':run,'page':Paginator(rows,30).get_page(request.GET.get('page'))})
def company_detail(request,run_id,company_id):
    row=get_object_or_404(CompanyResult,run_id=run_id,company_id=company_id)
    if request.method=='POST':
        form=ManualOverrideForm(request.POST)
        if form.is_valid(): row.manual_override_class=form.cleaned_data['manual_override_class'];row.manual_override_comment=form.cleaned_data['manual_override_comment'];row.manual_override_at=timezone.now();row.save();return redirect(request.path)
    else: form=ManualOverrideForm(initial={'manual_override_class':row.manual_override_class or row.final_class,'manual_override_comment':row.manual_override_comment})
    evidence=row.evidence or {}
    return render(request,'classifier_ui/company_detail.html',{
        'row':row,'run':row.run,'form':form,
        'rule_fired':evidence.get('rule_fired',[]),
        'counter_evidence':evidence.get('counter_evidence',[]),
        'representative_operations':evidence.get('representative_operations',[]),
        'website_search':evidence.get('website_search',{}),
        'site_keywords':evidence.get('site_keywords',[]),
        'site_keyphrases':evidence.get('site_keyphrases',[]),
        'site_signals':evidence.get('site_signals',[]),
    })
def download(request,run_id,kind):
    run=get_object_or_404(ClassificationRun,id=run_id)
    allowed={'excel':'classification_result.xlsx','csv':'company_results.csv','json':'company_results.json'}
    if kind not in allowed or run.status!='COMPLETED': raise Http404()
    path=Path(run.output_directory)/'output'/allowed[kind]
    if not path.is_file(): raise Http404()
    overrides={item.company_id:item.effective_final_class for item in run.companies.exclude(manual_override_class__isnull=True).exclude(manual_override_class='')}
    if not overrides: return FileResponse(path.open('rb'),as_attachment=True,filename=allowed[kind])
    if kind=='json':
        rows=json.loads(path.read_text(encoding='utf-8'))
        for row in rows:
            if str(row.get('company_id')) in overrides:
                row['original_final_class']=row.get('final_class');row['final_class']=overrides[str(row['company_id'])];row['manual_override_applied']=True
        return HttpResponse(json.dumps(rows,ensure_ascii=False,indent=2,default=str),content_type='application/json',headers={'Content-Disposition':'attachment; filename=company_results.json'})
    if kind=='csv':
        import pandas as pd
        data=pd.read_csv(path,dtype=str);data['original_final_class']=data.get('final_class','')
        data['final_class']=[overrides.get(str(cid),value) for cid,value in zip(data['company_id'],data['final_class'])];data['manual_override_applied']=data['company_id'].astype(str).isin(overrides)
        return HttpResponse(data.to_csv(index=False),content_type='text/csv; charset=utf-8',headers={'Content-Disposition':'attachment; filename=company_results.csv'})
    from openpyxl import load_workbook
    book=load_workbook(path);sheet=book['Классификация_без_ответов'];headers={str(c.value):c.column for c in sheet[1]}
    if 'original_final_class' not in headers: headers['original_final_class']=sheet.max_column+1;sheet.cell(1,headers['original_final_class']).value='original_final_class'
    if 'manual_override_applied' not in headers: headers['manual_override_applied']=sheet.max_column+1;sheet.cell(1,headers['manual_override_applied']).value='manual_override_applied'
    for index in range(2,sheet.max_row+1):
        cid=str(sheet.cell(index,headers['company_id']).value)
        if cid in overrides:
            sheet.cell(index,headers['original_final_class']).value=sheet.cell(index,headers['final_class']).value;sheet.cell(index,headers['final_class']).value=overrides[cid];sheet.cell(index,headers['manual_override_applied']).value=True
    stream=io.BytesIO();book.save(stream);stream.seek(0)
    return FileResponse(stream,as_attachment=True,filename='classification_result.xlsx')
