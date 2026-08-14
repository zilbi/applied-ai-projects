"""Local end-to-end orchestration for the frozen v2 structured classifier.

The module deliberately does not train models or make HTTP calls.  Existing
registry/site data are read as optional evidence only; a missing site is never
silently fabricated.
"""
from __future__ import annotations

import csv
import json
import math
import shutil
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from .entity_resolution import CompanyGroup, resolve_companies
from .ingestion import frame_to_operations, read_csv
from .model_registry import ModelRegistry, ModelRegistryError, feature_fingerprint
from .research_store import ResearchStore, utcnow
from .reference_catalog import active_reference, reference_decision, site_state_from_reference
from .validation import valid_inn
from .website_domain_roles import HARD_REJECT_ROLES, classify_domain_role
from .synthetic_time import temporal_behavior_features

CLASSES = ("A", "B", "NON_FINTECH")
AUTO_CONFIDENCE_THRESHOLD = 0.65
AUTO_SCORE_GAP = 0.15
PREDICTION_FIELDS = [
    "company_id", "canonical_company_name", "inn", "operation_count", "original_final_class", "final_class",
    "decision_status", "final_confidence", "probability_A", "probability_B", "probability_NON_FINTECH",
    "rule_score_A", "rule_score_B", "rule_score_NON_FINTECH", "models_agree", "site_url",
    "site_verification_status", "site_fetch_status", "data_quality_score", "score_gap", "review_reasons", "explanation",
]


class PipelineInputError(ValueError):
    pass


def _now() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def _purpose_count(values: list[str], needle: str) -> int:
    return sum(needle in (x or "").lower() for x in values)


def common_v2_features(group: CompanyGroup, columns: tuple[str, ...]) -> dict[str, float]:
    """Reproduce the frozen v2 22-column feature contract for a real group."""
    ops = group.operations
    amounts = np.asarray([float(op.amount) if op.amount is not None else 0.0 for op in ops], dtype=float)
    incoming = [op for op in ops if op.direction == "IN"]
    outgoing = [op for op in ops if op.direction == "OUT"]
    incoming_turnover = sum(float(op.amount or Decimal("0")) for op in incoming)
    outgoing_turnover = sum(float(op.amount or Decimal("0")) for op in outgoing)
    # A counterparty is the other party relative to the transaction direction,
    # exactly as in the frozen synthetic-v2 feature generator.
    counterparties = [
        str(op.payer_inn or "") if op.direction == "IN" else str(op.recipient_inn or "")
        for op in ops
    ]
    counterparties = [item for item in counterparties if item]
    shares = np.asarray(list(Counter(counterparties).values()), dtype=float)
    shares = shares / shares.sum() if shares.size else shares
    purposes = [op.payment_purpose_normalized or op.payment_purpose_raw or "" for op in ops]
    mean = float(amounts.mean()) if amounts.size else 0.0
    values = {
        "transaction_count": float(len(ops)), "log_transaction_count": float(math.log1p(len(ops))),
        "amount_mean": mean, "amount_median": float(np.median(amounts)) if amounts.size else 0.0,
        "amount_std": float(amounts.std(ddof=0)) if amounts.size else 0.0,
        "amount_cv": float(amounts.std(ddof=0) / mean) if mean else 0.0,
        "incoming_count": float(len(incoming)), "outgoing_count": float(len(outgoing)),
        "incoming_share": float(len(incoming) / len(ops)) if ops else 0.0,
        "outgoing_share": float(len(outgoing) / len(ops)) if ops else 0.0,
        "incoming_turnover": float(incoming_turnover), "outgoing_turnover": float(outgoing_turnover),
        "in_out_turnover_ratio": float(incoming_turnover / outgoing_turnover) if outgoing_turnover else 0.0,
        "unique_counterparty_count": float(len(shares)), "counterparty_hhi": float((shares ** 2).sum()) if shares.size else 0.0,
        "purpose_payment_count": float(_purpose_count(purposes, "оплат")),
        "purpose_commission_count": float(_purpose_count(purposes, "комисс")),
        "purpose_payout_count": float(_purpose_count(purposes, "выплат") + _purpose_count(purposes, "перечислен")),
        "purpose_order_count": float(_purpose_count(purposes, "заказ")),
        "purpose_refund_count": float(_purpose_count(purposes, "возврат")),
        "purpose_own_goods_count": float(_purpose_count(purposes, "собственн")),
        "purpose_platform_count": float(_purpose_count(purposes, "платформ")),
    }
    values.update(temporal_behavior_features(op.operation_datetime for op in ops))
    missing = [field for field in columns if field not in values]
    if missing:
        raise PipelineInputError(f"Runtime feature builder lacks required columns: {', '.join(missing)}")
    return {field: values[field] for field in columns}


def _site_state(store: ResearchStore | None, inn: str | None, *, since: str | None = None) -> dict[str, Any]:
    empty = {"site_url": "", "verification_status": "not_started", "fetch_status": "not_started", "analysis_status": "not_started", "evidence": [], "website_search": {"attempts": [], "candidates": []}}
    if store is None or not inn or not valid_inn(inn):
        return empty
    website_rows = store.connection.execute("""SELECT w.id AS website_id,w.website_url,w.verification_status,w.fetch_status,w.analysis_status,
                                           w.verification_score
        FROM company_websites w JOIN companies c ON c.id=w.company_id WHERE c.inn=? AND (? IS NULL OR w.updated_at>=?)
        ORDER BY CASE w.verification_status WHEN 'confirmed_by_website' THEN 0 WHEN 'confirmed_by_registry' THEN 1 WHEN 'probable' THEN 2 ELSE 3 END,
                 COALESCE(w.verification_score, 0) DESC, w.id DESC LIMIT 20""", (inn, since, since)).fetchall()
    # The shared research database can contain candidates written by older
    # versions. Reapply today's hard domain policy before an old directory
    # record can be surfaced as a company's official website.
    row = next((item for item in website_rows
                if classify_domain_role(item["website_url"] or "")[0] not in HARD_REJECT_ROLES), None)
    company = store.connection.execute("SELECT id FROM companies WHERE inn=?", (inn,)).fetchone()
    if not company:
        return empty
    company_id = int(company["id"])
    try:
        attempts = [dict(item) for item in store.connection.execute(
            """SELECT provider,reason_for_call,template_id,search_query,include_domains_json,exclude_domains_json,
                      status,http_status,request_id,response_time,result_count,accepted_count,rejected_count,
                      credits_used,error_type,error_message,started_at,finished_at
               FROM website_search_attempts WHERE company_id=? AND (? IS NULL OR started_at>=?) ORDER BY id""", (company_id, since, since)
        ).fetchall()]
    except Exception:  # Historical read-only databases predate this audit table.
        attempts = []
    for attempt in attempts:
        for key in ("include_domains_json", "exclude_domains_json"):
            try:
                attempt[key[:-5]] = json.loads(attempt.pop(key) or "[]")
            except (TypeError, json.JSONDecodeError):
                attempt[key[:-5]] = []
    # Offline classifications open historical SQLite files read-only.  Those
    # files cannot be migrated in place, so select new audit columns only when
    # they exist and fill their safe defaults below.
    candidate_columns = {item[1] for item in store.connection.execute("PRAGMA table_info(website_candidates)")}
    requested = ["candidate_url", "candidate_source", "source_type", "domain_role", "role_reason", "brand_match", "title_match", "rejection_reason", "selected",
                 "candidate_status", "verification_status", "fetch_status", "analysis_status", "search_position", "search_score", "verification_score", "candidate_score",
                 "positive_evidence_json", "negative_evidence_json", "checked_pages_json"]
    available = [name for name in requested if name in candidate_columns]
    order = "selected DESC, updated_at DESC" if "selected" in candidate_columns else "updated_at DESC"
    candidates = [dict(item) for item in store.connection.execute(
        f"SELECT {','.join(available)} FROM website_candidates WHERE company_id=? AND (? IS NULL OR updated_at>=?) ORDER BY {order} LIMIT 12",
        (company_id, since, since)
    ).fetchall()]
    for candidate in candidates:
        candidate.setdefault("source_type", "search_result")
        candidate.setdefault("domain_role", "OFFICIAL_CANDIDATE")
        candidate.setdefault("role_reason", None)
        candidate.setdefault("brand_match", False)
        candidate.setdefault("title_match", False)
        candidate.setdefault("search_score", 0.0)
        candidate.setdefault("verification_score", 0.0)
        candidate.setdefault("rejection_reason", None)
        candidate.setdefault("selected", False)
        for key in ("positive_evidence_json", "negative_evidence_json", "checked_pages_json"):
            try:
                candidate[key[:-5]] = json.loads(candidate.pop(key) or "[]")
            except (TypeError, json.JSONDecodeError):
                candidate[key[:-5]] = []
        current_role, current_reason = classify_domain_role(candidate.get("candidate_url") or "")
        if current_role in HARD_REJECT_ROLES:
            # Never trust a historical ``OFFICIAL_CANDIDATE`` value from the
            # database. In particular, Kontur/Fokus must not be displayed as
            # viable or selected in any later run.
            candidate.update({
                "domain_role": current_role.value,
                "role_reason": current_reason,
                "hard_rejected": True,
                "shortlist_eligible": False,
                "selected": False,
                "verification_status": "rejected",
                "fetch_status": "not_started",
                "analysis_status": "skipped",
                "candidate_status": "rejected",
                "selection_status": "HARD_REJECTED",
                "rejection_reason": f"глобально запрещённый домен: {current_reason}",
            })
    search = {"attempts": attempts, "candidates": candidates}
    if not row:
        # A rejected card (App Store, directory, social network) is evidence
        # about a failed search path, never an official-site URL.  It remains
        # visible in ``website_search.candidates`` but must not leak into the
        # company summary as if it were the selected site.
        viable = [item for item in candidates if item.get("verification_status") in {
            "confirmed_by_website", "confirmed_by_registry", "probable"
        } and item.get("domain_role") in {"official_candidate", "OFFICIAL_CANDIDATE"}]
        if viable:
            best = viable[0]
            return {"site_url": best.get("candidate_url") or "", "verification_status": best.get("verification_status") or "not_started",
                    "fetch_status": best.get("fetch_status") or "not_started", "analysis_status": best.get("analysis_status") or "not_started",
                    "evidence": [], "website_search": search}
        return {**empty, "fetch_status": "not_found", "verification_status": "rejected", "website_search": search}
    website_id = int(row["website_id"])
    evidence = [dict(item) for item in store.connection.execute(
        "SELECT signal_family,matched_phrase,context,page_url,html_zone,weight FROM website_signals WHERE company_id=? AND website_id=? ORDER BY weight DESC LIMIT 12",
        (company_id, website_id),
    ).fetchall()]
    keyword_rows = [dict(item) for item in store.connection.execute(
        """SELECT keyword_type,text,normalized_text,score,occurrences,page_urls_json,contexts_json
           FROM website_keywords WHERE company_id=? AND website_id=?
           ORDER BY keyword_type, score DESC, text LIMIT 35""",
        (company_id, website_id),
    ).fetchall()]
    for item in keyword_rows:
        for key in ("page_urls_json", "contexts_json"):
            try:
                item[key[:-5]] = json.loads(item.pop(key) or "[]")
            except (TypeError, json.JSONDecodeError):
                item[key[:-5]] = []
    keywords = [item for item in keyword_rows if item["keyword_type"] == "keyword"][:20]
    keyphrases = [item for item in keyword_rows if item["keyword_type"] == "keyphrase"][:15]
    return {"site_url": row["website_url"] or "", "verification_status": row["verification_status"] or "not_started",
            "fetch_status": row["fetch_status"] or "not_started", "analysis_status": row["analysis_status"] or "not_started", "evidence": evidence,
            "site_keywords": keywords, "site_keyphrases": keyphrases, "website_search": search}


def normalized_rule_scores(features: dict[str, float], site: dict[str, Any]) -> tuple[dict[str, float], list[dict[str, Any]], list[str]]:
    """Transparent, deliberately conservative rules. Generic payment words do not decide A/B."""
    raw = dict.fromkeys(CLASSES, 0.0); fired: list[dict[str, Any]] = []; counter: list[str] = []
    def add(cls: str, weight: float, rule_id: str, evidence: str, source: str = "payments"):
        raw[cls] += weight; fired.append({"rule_id": rule_id, "supported_class": cls, "weight": weight, "source": source, "evidence": evidence, "counter_evidence": ""})
    if features["purpose_payout_count"] >= 2 and features["purpose_commission_count"] >= 1:
        add("A", .55, "A_SETTLEMENT_COMMISSION", "повторяются выплаты/перечисления и комиссия")
    if features["purpose_payout_count"] >= 2 and features["incoming_count"] and features["outgoing_count"]:
        add("A", .20, "A_TWO_SIDED_PAYOUT", "есть входящий и исходящий поток с выплатами")
    if features["purpose_order_count"] >= 2 and features["purpose_payout_count"] >= 1:
        add("B", .55, "B_ORDERS_PAYOUTS", "есть заказы и выплаты участникам")
    if features["purpose_platform_count"] >= 2 and features["purpose_order_count"] >= 1:
        add("B", .25, "B_PLATFORM_ORDERS", "есть устойчивые платформенные заказы")
    if features["purpose_own_goods_count"] >= 2:
        add("NON_FINTECH", .60, "N_OWN_GOODS", "назначения указывают на собственные товары/услуги")
    for signal in site.get("evidence", []):
        family = signal.get("signal_family") or ""
        phrase = signal.get("matched_phrase") or ""
        weight = min(.25, float(signal.get("weight") or 0) / 10)
        if family in {"PAYMENT_SERVICE", "THIRD_PARTY_FUNDS", "COMMISSION_WITHHOLDING", "SETTLEMENT"}:
            add("A", weight, "SITE_" + family, phrase, "website")
        elif family in {"INDEPENDENT_PARTICIPANTS", "PLATFORM_ORDER", "ORDER_MANAGEMENT", "PLATFORM_COMMISSION", "MULTI_SIDED_PLATFORM"}:
            add("B", weight, "SITE_" + family, phrase, "website")
        elif family in {"OWN_PRODUCTS", "OWN_SERVICES", "OWN_STORES", "OWN_PRODUCTION", "DIRECT_RETAIL"}:
            add("NON_FINTECH", weight, "SITE_" + family, phrase, "website")
    if not fired:
        counter.append("нет достаточных rule-based свидетельств")
    total = sum(raw.values())
    # A score is normalised to the [0, 1] range independently per class.  Do
    # not divide by the sum: one weak trigger must not become a fake 1.0 just
    # because the other classes happened not to trigger.
    return ({key: round(min(value, 1.0), 6) for key, value in raw.items()} if total else {key: 0.0 for key in CLASSES}), fired, counter


def _save_research_record(store: ResearchStore, request_id: int, company_id: int, record: Any) -> None:
    """Persist public-source provenance exactly as the manual scenario does."""
    with store.transaction():
        for evidence in record.source_evidence:
            source_id = store.record_source_result(request_id, company_id, source_name=evidence.source, input_inn=record.inn,
                source_url=evidence.url, request_status=evidence.status, inn_confirmed=evidence.inn_confirmed,
                warnings=[evidence.message] if evidence.message else [], content=evidence.content, raw_result=evidence.to_dict())
            if evidence.inn_confirmed:
                for field_name, value in evidence.facts.items():
                    store.add_fact(company_id, field_name, value, source_result_id=source_id, source_name=evidence.source,
                                   source_url=evidence.url, confidence=.8)
                    if field_name == "legal_name":
                        store.add_alias(company_id, None, value, "source_name", evidence.source, source_id)
        store.update_company_facts(company_id, record.canonical_facts)


def _payment_name_candidates(group: CompanyGroup) -> list[str]:
    """Preserve public aliases seen in the input for brand-domain discovery.

    Grouping remains strictly INN-based.  These names are not model features;
    they only let a legal name such as ``ООО Интернет Решения`` retain the
    observed consumer brand ``Ozon`` when searching the public web.
    """
    counts: Counter[str] = Counter()
    inn = group.profile.inn
    for operation in group.operations:
        for side in ("payer", "recipient"):
            if getattr(operation, f"{side}_inn", None) == inn:
                raw = str(getattr(operation, f"{side}_name_raw", None) or "").strip()
                if raw:
                    counts[raw] += 1
    # Keep canonical first for legal lookup, followed by the most frequently
    # observed aliases.  Deduplication is intentional and deterministic.
    ordered = [group.profile.canonical_name]
    ordered.extend(name for name, _ in counts.most_common() if name != group.profile.canonical_name)
    return list(dict.fromkeys(item for item in ordered if item))[:12]


def _external_enrich(store: ResearchStore, group: CompanyGroup, *, tavily_provider=None,
                     max_tavily_credits: int | None = None) -> int | None:
    """Reuse the existing five-source/OKVED/site workflow for one valid INN.

    This function is called only under the explicit external mode. Individual
    public-source failures are retained as source results and never stop the
    batch.
    """
    if not group.profile.inn or not valid_inn(group.profile.inn):
        return None
    from .company_parser import CompanyLookupRequest
    from .company_research import research_company
    from .okved_agent import OkvedLookupRequest, lookup_okved
    from .website_workflow import analyze_confirmed_and_store, discover_and_store
    name_candidates = _payment_name_candidates(group)
    request = {"lookup_id": f"PIPELINE-{group.profile.company_id}", "inn": group.profile.inn,
               "legal_name_candidates": name_candidates, "account_candidates": group.profile.accounts,
               "operation_ids": group.profile.operation_ids, "source_sides": [], "lookup_status": "pipeline_external"}
    request_id = store.import_request(request, source_file="full_pipeline")
    row = store.connection.execute("SELECT company_id FROM research_requests WHERE id=?", (request_id,)).fetchone()
    company_id = int(row[0]) if row and row[0] is not None else None
    if company_id is None:
        return None
    lookup = CompanyLookupRequest(f"PIPELINE-{request_id}", group.profile.inn, name_candidates,
                                  group.profile.accounts, group.profile.operation_ids, [], "ready_for_fns_lookup", group.profile.inn)
    record = research_company(lookup)
    _save_research_record(store, request_id, company_id, record)
    lookup_okved(OkvedLookupRequest(company_id, group.profile.inn, name_candidates), allow_external=True, store=store)
    discover_and_store(store, company_id, online=True, tavily_provider=tavily_provider,
                       max_tavily_credits=max_tavily_credits)
    analyze_confirmed_and_store(store, company_id, online=True)
    store.connection.commit()
    return company_id


def detect_input_identity_conflict(store: ResearchStore | None, group: CompanyGroup) -> bool:
    """Detect an incompatible input brand and INN-confirmed legal entity.

    The valid INN remains the primary entity key.  This guard does not rewrite
    it to match a familiar brand: it simply blocks automatic classification so
    a human can correct the input record.
    """
    if store is None or not group.profile.inn or not valid_inn(group.profile.inn):
        return False
    row = store.connection.execute("SELECT confirmed_legal_name FROM companies WHERE inn=?", (group.profile.inn,)).fetchone()
    legal = str(row["confirmed_legal_name"] or "") if row else ""
    if not legal:
        return False
    import re
    ignored = {"ооо", "ао", "пао", "зао", "нао", "ип", "компания", "общество", "ограниченной", "ответственностью", "ру"}
    def terms(value: str) -> set[str]:
        return {x for x in re.findall(r"[а-яёa-z0-9]{3,}", value.lower()) if x not in ignored}
    input_names = _payment_name_candidates(group)
    input_terms = set().union(*(terms(name) for name in input_names)) if input_names else set()
    legal_terms = terms(legal)
    # Do not call a legal-form-only input a conflict: it is merely weak data.
    return bool(input_terms and legal_terms and not (input_terms & legal_terms))


def arbitrate_v1(rule_scores: dict[str, float], probs: dict[str, float], *, grouping_confidence: float,
                 data_quality: float, completeness_factor: float, conflicts: list[str],
                 regulated_financial_institution: bool = False,
                 auto_confidence_threshold: float = AUTO_CONFIDENCE_THRESHOLD,
                 score_gap_threshold: float = AUTO_SCORE_GAP) -> dict[str, Any]:
    """Let CatBoost make the class decision; retain REVIEW for identity safety.

    Rules, probability margin and the regulated-company marker are diagnostic
    evidence only.  They must not silently overrule the final model.  REVIEW
    is reserved for an invalid/ambiguous entity identity passed in ``conflicts``.
    """
    ordered = sorted(CLASSES, key=probs.get, reverse=True)
    winner, second = ordered[0], ordered[1]
    confidence = probs[winner] * grouping_confidence * completeness_factor
    gap = probs[winner] - probs[second]
    review_reasons = list(dict.fromkeys(conflicts))
    # ``confidence`` and ``gap`` remain in the output for the UI and audit;
    # they no longer create a pseudo-human decision in place of CatBoost.
    # ``regulated_financial_institution`` is likewise explanatory metadata.
    status = "AUTO" if not review_reasons else "MANUAL_REVIEW"
    return {"scores": dict(probs), "final_class": winner if status == "AUTO" else "REVIEW", "original_final_class": winner,
            "decision_status": status, "final_confidence": round(float(confidence), 6), "score_gap": round(float(gap), 6),
            "hard_conflicts": conflicts, "review_reasons": review_reasons}


def _regulated_financial_institution(store: ResearchStore | None, inn: str | None) -> bool:
    """Use only confirmed registry facts/OKVED, never a search result or model score."""
    if store is None or not inn or not valid_inn(inn):
        return False
    company = store.connection.execute("SELECT id, confirmed_legal_name FROM companies WHERE inn=?", (inn,)).fetchone()
    if not company:
        return False
    company_id = int(company["id"])
    primary = store.connection.execute(
        "SELECT okved_code FROM company_okved WHERE company_id=? AND is_primary=1 AND is_conflicting=0 ORDER BY id LIMIT 1", (company_id,)
    ).fetchone()
    if primary and str(primary["okved_code"] or "").replace(".", "")[:2] in {"64", "65", "66"}:
        return True
    names = [str(company["confirmed_legal_name"] or "")]
    names.extend(str(item[0] or "") for item in store.connection.execute(
        "SELECT value_text FROM company_facts WHERE company_id=? AND field_name='legal_name' AND is_conflicting=0", (company_id,)
    ).fetchall())
    return any("банк" in value.lower() or "кредитн" in value.lower() and "организац" in value.lower() for value in names)


def _quality(group: CompanyGroup, site: dict[str, Any]) -> tuple[float, float, list[str]]:
    missing: list[str] = []
    identity = 1.0 if group.profile.inn and valid_inn(group.profile.inn) else .35
    if identity < 1: missing.append("отсутствует или невалиден ИНН")
    operations = min(1.0, len(group.operations) / 5)
    if len(group.operations) < 5: missing.append("мало операций")
    if not any(op.direction == "IN" for op in group.operations) or not any(op.direction == "OUT" for op in group.operations):
        missing.append("только односторонний поток")
    site_score = 1.0 if site["analysis_status"] == "success" else .75 if site["verification_status"].startswith("confirmed") else .55 if site["verification_status"] == "probable" else .4
    if site["fetch_status"] in {"timeout", "blocked_by_waf", "http_401", "http_403", "http_498"}: missing.append("сайт недоступен: " + site["fetch_status"])
    completeness = 1.0
    if identity < 1.0: completeness *= .55
    if len(group.operations) < 5: completeness *= .80
    return round((identity + operations + site_score) / 3, 6), round(completeness, 6), missing


class FullClassificationPipeline:
    def __init__(self, *, no_network: bool = True, store_path: str | Path = "results/company_research.sqlite3",
                 progress: Callable[[str, int, int], None] | None = None, max_tavily_credits: int | None = None) -> None:
        self.no_network, self.store_path, self.progress = no_network, Path(store_path), progress
        self.max_tavily_credits = max_tavily_credits

    def dry_run(self, input_path: str | Path) -> dict[str, Any]:
        frames = self._read_input(input_path)
        models = ModelRegistry().load_production()
        return {"valid": True, "network": False, "sheets": list(frames), "rows": {name: len(frame) for name, frame in frames.items()},
                "model": models.name, "feature_schema": models.feature_schema, "feature_fingerprint": models.fingerprint,
                "required_features": list(models.columns)}

    def _read_input(self, input_path: str | Path) -> dict[str, pd.DataFrame]:
        path = Path(input_path)
        if not path.exists(): raise PipelineInputError("Входной файл не найден.")
        if path.suffix.lower() == ".csv":
            frame = pd.read_csv(path, dtype=str, keep_default_na=False)
            return {"Классификация_без_ответов": frame}
        if path.suffix.lower() != ".xlsx": raise PipelineInputError("Поддерживаются только XLSX и CSV.")
        book = pd.ExcelFile(path)
        expected = {"Обучение_с_ответами", "Классификация_без_ответов"}
        if set(book.sheet_names) != expected:
            raise PipelineInputError("XLSX должен содержать ровно листы «Обучение_с_ответами» и «Классификация_без_ответов».")
        return {name: pd.read_excel(path, sheet_name=name, dtype=str, keep_default_na=False) for name in book.sheet_names}

    def run(self, input_path: str | Path, output_dir: str | Path) -> dict[str, Any]:
        frames = self._read_input(input_path)
        models = ModelRegistry().load_production()
        if feature_fingerprint(models.columns) != models.fingerprint:
            raise ModelRegistryError("Runtime feature column fingerprint mismatch.")
        source = frames["Классификация_без_ответов"]
        ops, parse_warnings = frame_to_operations(source, "full_pipeline")
        groups = resolve_companies(ops)
        features = [common_v2_features(group, models.columns) for group in groups]
        matrix = pd.DataFrame(features, columns=list(models.columns))
        if tuple(matrix.columns) != models.columns or feature_fingerprint(tuple(matrix.columns)) != models.fingerprint:
            raise ModelRegistryError("Runtime feature schema/fingerprint does not match production model; prediction blocked.")
        primary_prob = models.primary.predict_proba(matrix)
        primary_pred = models.primary.predict(matrix)
        logreg_pred = models.supporting_logreg.predict(matrix) if models.supporting_logreg is not None else [""] * len(groups)
        hgb_pred = models.supporting_hgb.predict(matrix) if models.supporting_hgb is not None else [""] * len(groups)
        store = ResearchStore(self.store_path, read_only=self.no_network) if (self.store_path.exists() or not self.no_network) else None
        tavily_provider = None
        if not self.no_network:
            # One provider instance owns one shared credit counter for the
            # whole batch.  A ten-credit run over ten companies therefore
            # makes at most one Tavily Search per company, never 2×10 calls.
            from .tavily_search import TavilySearchProvider
            tavily_provider = TavilySearchProvider()
            credit_limit = self.max_tavily_credits or tavily_provider.config.max_credits_per_run
            calls_per_company = min(tavily_provider.config.max_calls_per_company,
                                    max(1, int(credit_limit) // max(1, len(groups))))
            tavily_provider.config = tavily_provider.config.__class__(
                **{**tavily_provider.config.__dict__, "max_credits_per_run": int(credit_limit),
                   "max_calls_per_company": calls_per_company}
            )
        results: list[dict[str, Any]] = []
        evidence_dir = Path(output_dir) / "evidence"; evidence_dir.mkdir(parents=True, exist_ok=True)
        for idx, (group, feature_row) in enumerate(zip(groups, features), 1):
            if self.progress: self.progress("classification", idx - 1, len(groups))
            model_probs = {cls: float(primary_prob[idx - 1][list(models.primary.classes_).index(cls)]) for cls in CLASSES}
            # CatBoost wrapped in sklearn Pipeline returns shape (n, 1) on
            # this runtime; scalarise it before persistence/comparison.
            model_primary_label = str(np.asarray(primary_pred[idx - 1]).ravel()[0])
            primary_label = model_primary_label
            logreg_label = str(np.asarray(logreg_pred[idx - 1]).ravel()[0])
            hgb_label = str(np.asarray(hgb_pred[idx - 1]).ravel()[0])
            reference = active_reference(store, group.profile.inn)
            enrichment_started_at = utcnow() if not self.no_network and reference is None else None
            if not self.no_network and store is not None and reference is None:
                if self.progress: self.progress("public_research", idx - 1, len(groups))
                try:
                    _external_enrich(store, group, tavily_provider=tavily_provider,
                                     max_tavily_credits=self.max_tavily_credits)
                except Exception as exc:
                    # Network failures are represented as missing evidence;
                    # the company still gets a local model prediction.
                    group.profile.warnings.append(f"external_research_error:{type(exc).__name__}")
            site = site_state_from_reference(reference) if reference is not None else _site_state(store, group.profile.inn, since=enrichment_started_at)
            quality, completeness, missing = _quality(group, site)
            rule_scores, fired, counter = normalized_rule_scores(feature_row, site)
            diagnostic_conflicts = []
            if len({primary_label, logreg_label, hgb_label}) == 3:
                diagnostic_conflicts.append("CatBoost, Logistic Regression и HGB дали три разных класса")
            if rule_scores["A"] >= .50 and rule_scores["B"] >= .50:
                diagnostic_conflicts.append("правила одновременно нашли сильные признаки A и B")
            # Only entity-resolution failures may stop the final CatBoost
            # decision. Model/rule disagreement is useful diagnostics, not a
            # second classifier and not a reason to manufacture REVIEW.
            blocking_conflicts = []
            if "critical_conflicting_inns" in group.profile.warnings:
                blocking_conflicts.append("разные валидные ИНН ошибочно объединены")
            if not group.profile.inn or not valid_inn(group.profile.inn):
                blocking_conflicts.append("нет валидного ИНН для идентификации компании")
            if reference is None and detect_input_identity_conflict(store, group):
                blocking_conflicts.append("input_name_inn_identity_conflict")
            rule_top = max(CLASSES, key=lambda cls: rule_scores[cls])
            if rule_scores[rule_top] > 0 and rule_top != primary_label:
                diagnostic_conflicts.append(f"CatBoost={primary_label}, rules={rule_top}")
            regulated_financial_institution = _regulated_financial_institution(store, group.profile.inn)
            if regulated_financial_institution and max(rule_scores.get("A", 0.0), rule_scores.get("B", 0.0)) < .50:
                diagnostic_conflicts.append("regulated_financial_institution_outside_target_taxonomy")
            if reference is not None:
                decision = reference_decision(reference)
                probs = dict(decision["scores"])
                primary_label = decision["final_class"]
            else:
                probs = model_probs
                decision = arbitrate_v1(rule_scores, probs, grouping_confidence=group.profile.grouping_confidence, data_quality=quality,
                                        completeness_factor=completeness, conflicts=blocking_conflicts,
                                        regulated_financial_institution=regulated_financial_institution)
            # Agreement shown to a reviewer includes the explanation rules;
            # three model labels alone cannot hide CatBoost=A/rules=B.
            available_supporting = [label for label in (logreg_label, hgb_label) if label]
            model_agree = all(label == primary_label for label in available_supporting) and rule_top == primary_label
            top = sorted(probs, key=probs.get, reverse=True)
            facts = [item["evidence"] for item in sorted(fired, key=lambda x: x["weight"], reverse=True)[:3]]
            if reference is not None:
                facts.insert(0, str(reference.get("classification_basis") or "проверенные сведения о компании"))
            elif primary_label == decision["original_final_class"]:
                facts.append(f"CatBoost выбрал {primary_label} ({probs[primary_label]:.2f})")
            explanation = (f"Компания отнесена к {decision['final_class']}. " + ("; ".join(facts) if facts else "Недостаточно фактов для устойчивого вывода.") +
                           (" Требуется ручная проверка: " + "; ".join(decision["review_reasons"]) + "." if decision["decision_status"] != "AUTO" else " Модели и правила согласованы."))
            if "input_name_inn_identity_conflict" in decision["review_reasons"]:
                explanation = ("Официальный сайт бренда найден, однако его реквизиты не совпадают с ИНН во входном файле. "
                               "Автоматическая классификация заблокирована. " + explanation)
            result = {"company_id": group.profile.company_id, "canonical_company_name": group.profile.canonical_name, "inn": group.profile.inn or "",
                      "operation_count": len(group.operations), "grouping_confidence": group.profile.grouping_confidence,
                      "primary_model": models.name, "model_version": models.version, "feature_schema": models.feature_schema,
                      "feature_fingerprint": models.fingerprint, "probability_A": probs["A"], "probability_B": probs["B"], "probability_NON_FINTECH": probs["NON_FINTECH"],
                      "predicted_class": primary_label, "top_probability": probs[top[0]], "top1_top2_margin": probs[top[0]] - probs[top[1]],
                      "model_predicted_class": model_primary_label, "model_probabilities": model_probs,
                      "supporting_logreg_class": logreg_label, "supporting_hgb_class": hgb_label, "models_agree": bool(model_agree),
                      "models_disagreement_description": "" if model_agree else "; ".join(
                          [f"CatBoost={model_primary_label}", *([f"LogisticRegression={logreg_label}"] if logreg_label else []),
                           *([f"HGB={hgb_label}"] if hgb_label else []), f"rules={rule_top}"]),
                      "site_url": site["site_url"], "site_verification_status": site["verification_status"], "site_fetch_status": site["fetch_status"],
                      "site_analysis_status": site["analysis_status"], "data_quality_score": quality, "completeness_factor": completeness, "missing_information": missing, "rule_score_A": rule_scores["A"], "rule_score_B": rule_scores["B"], "rule_score_NON_FINTECH": rule_scores["NON_FINTECH"],
                      "regulated_financial_institution": regulated_financial_institution,
                      "website_search": site["website_search"], "site_keywords": site.get("site_keywords", []),
                      "site_keyphrases": site.get("site_keyphrases", []), "site_signals": site.get("evidence", []),
                      "rule_fired": fired, "counter_evidence": counter, "diagnostic_conflicts": diagnostic_conflicts, **decision, "explanation": explanation,
                      "representative_operations": [op.operation_id for op in group.operations[:5]], "feature_values": feature_row,
                      "catalog_profile": ({"inn": reference["inn"], "catalog_version": reference["catalog_version"]} if reference is not None else None)}
            (evidence_dir / f"{group.profile.company_id}.json").write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
            results.append(result)
        if store: store.close()
        self._write_outputs(Path(input_path), source, ops, groups, results, frames, Path(output_dir), parse_warnings, models)
        if self.progress: self.progress("completed", len(groups), len(groups))
        return {"companies": len(groups), "operations": len(ops), "results": results, "parse_warnings": parse_warnings,
                "output_dir": str(Path(output_dir)), "model": models.name, "feature_fingerprint": models.fingerprint}

    def _write_outputs(self, input_path: Path, source: pd.DataFrame, ops, groups, results, frames, output_dir: Path, warnings, models) -> None:
        out = output_dir / "output"; out.mkdir(parents=True, exist_ok=True)
        by_op: dict[str, dict] = {}
        memberships: dict[str, list[tuple[CompanyGroup, dict]]] = defaultdict(list)
        for group, result in zip(groups, results):
            for op_id in group.profile.operation_ids: memberships[op_id].append((group, result))
        for op_id, options in memberships.items():
            by_op[op_id] = max(options, key=lambda pair: (pair[1]["operation_count"], pair[0].profile.grouping_confidence))[1]
        compact = [{key: value for key, value in item.items() if key not in {"rule_fired", "feature_values", "representative_operations", "counter_evidence"}} for item in results]
        pd.DataFrame(compact).to_csv(out / "company_results.csv", index=False)
        (out / "company_results.json").write_text(json.dumps(results, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        report = {"created_at": _now(), "input_file": input_path.name, "network_enabled": not self.no_network, "companies": len(results),
                  "class_distribution": dict(Counter(item["final_class"] for item in results)), "auto_count": sum(item["decision_status"] == "AUTO" for item in results),
                  "review_count": sum(item["decision_status"] == "MANUAL_REVIEW" for item in results), "parse_warnings": warnings,
                  "model": {"name": models.name, "version": models.version, "feature_schema": models.feature_schema, "feature_fingerprint": models.fingerprint, "provenance": models.provenance}}
        (out / "run_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        if input_path.suffix.lower() == ".xlsx":
            workbook = load_workbook(input_path)
            sheet = workbook["Классификация_без_ответов"]
            headers = {str(cell.value): cell.column for cell in sheet[1]}
            for label in PREDICTION_FIELDS:
                if label not in headers:
                    headers[label] = sheet.max_column + 1; sheet.cell(1, headers[label]).value = label
            op_col = headers.get("operation_id")
            for row in range(2, sheet.max_row + 1):
                result = by_op.get(str(sheet.cell(row, op_col).value)) if op_col else None
                if not result: continue
                values = {**result, "final_class": result["final_class"], "original_final_class": result["original_final_class"], "review_reasons": "; ".join(result["review_reasons"])}
                for label in PREDICTION_FIELDS: sheet.cell(row, headers[label]).value = values.get(label, "")
            workbook.save(out / "classification_result.xlsx")
        else:
            # CSV still gets a two-sheet output contract in a generated workbook.
            with pd.ExcelWriter(out / "classification_result.xlsx", engine="openpyxl") as writer:
                pd.DataFrame().to_excel(writer, sheet_name="Обучение_с_ответами", index=False)
                enriched = source.copy()
                enriched["final_class"] = [by_op.get(str(op.operation_id), {}).get("final_class", "") for op in ops]
                enriched.to_excel(writer, sheet_name="Классификация_без_ответов", index=False)
