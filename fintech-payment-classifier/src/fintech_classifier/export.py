from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


PREDICTION_COLUMNS = {
    "predicted_company_id": "company_id", "predicted_canonical_company_name": "canonical_name", "grouping_case": "grouping_case",
    "official_site": "official_site", "website_evidence": "website_evidence", "predicted_segment": "predicted_segment", "confidence": "confidence",
    "alternative_hypothesis": "alternative_hypothesis", "evidence_summary": "evidence_summary", "counter_evidence": "counter_evidence",
    "missing_information": "missing_information", "rationale": "rationale",
}


def export_predictions(input_path: str | Path, output_path: str | Path, predictions: dict[str, dict]) -> None:
    workbook = load_workbook(input_path)
    if "Классификация_без_ответов" not in workbook.sheetnames or len(workbook.sheetnames) != 2:
        raise ValueError("Ожидается исходный файл с ровно двумя листами и листом «Классификация_без_ответов»")
    sheet = workbook["Классификация_без_ответов"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    for row in range(2, sheet.max_row + 1):
        operation_id = str(sheet.cell(row, headers["operation_id"]).value)
        prediction = predictions.get(operation_id)
        if not prediction:
            continue
        for excel_column, value_key in PREDICTION_COLUMNS.items():
            column = headers.get(excel_column)
            if column:
                sheet.cell(row, column).value = prediction.get(value_key)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
